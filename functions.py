# 処理一覧
import re
import openpyxl


class GetRightNumber:
    def __init__(self):
        self.name = '右側から数字抽出(結果のみ)'
        self.explain = '結果のみを別シートに記入していく'

    def run(self, workbook, sheet_name, tgt_cols):
        """
        workbook 対象となるワークブック
        sheet name　処理もとのデータがあるシート
        tgt cols　処理を行う列
        """
        src_ws = workbook[sheet_name]  # コピー元のワークシート
        tgt_ws = workbook.create_sheet(title=self.name)  # コピー先のシート
        c = 1  # col 通し番号
        matcher = re.compile('\d+$')  # 数字
        for cols in tgt_cols:
            for col in cols:
                r = 1  # row　通し番号
                target_col = src_ws[col]
                for i in range(len(target_col)):
                    cell = src_ws[col][i].value  # col列のi行目
                    if cell is not None:
                        result = re.search(matcher, cell.rstrip())
                        if result:
                            # matchした場合
                            tgt_ws.cell(row=r, column=c).value = cell[:result.start()]
                            tgt_ws.cell(row=r, column=c + 1).value = int(result.group())
                        else:
                            # matchしない場合そのまま書き込み
                            tgt_ws.cell(row=r, column=c).value = cell
                    r += 1
                c += 2
            c += 1


class GetRightNumber2:
    def __init__(self):
        self.name = '右側から数字抽出'
        self.explain = '元のシートの値を別シートに書き込み、最後の列に結果を挿入していく'

    def run(self, workbook, sheet_name, tgt_cols):
        """
        workbook 対象となるワークブック
        sheet name　処理もとのデータがあるシート
        tgt cols　処理を行う列
        """
        src_ws = workbook[sheet_name]  # コピー元のワークシート
        # 元のシートの内容コピーして書き込み先の新しいシート作成
        tgt_ws = workbook.copy_worksheet(src_ws)  # コピー先のシート作成
        tgt_ws.title = self.name
        # 数字抜き出した部分を色付けて付け足す
        matcher = re.compile('\d+$')  # 数字抜き出すための正規表現
        c = tgt_ws.max_column + 2  # current col 現在書き込まれているセルの右端+αからスタート
        fill = openpyxl.styles.PatternFill(fgColor='FFFF00', bgColor="FFFF00", fill_type="solid")  # 色を塗るためのオブジェクト
        for cols in tgt_cols:
            for col in cols:
                r = 1  # row　通し番号
                target_col = src_ws[col]
                for i in range(len(target_col)):
                    cell = src_ws[col][i].value  # col列のi行目
                    if cell is not None:
                        result = re.search(matcher, cell.rstrip())
                        if result:
                            # matchした場合
                            tgt_ws.cell(row=r, column=c).value = cell[:result.start()]
                            tgt_ws.cell(row=r, column=c + 1).value = int(result.group())
                        else:
                            # matchしない場合そのまま書き込み
                            tgt_ws.cell(row=r, column=c).value = cell
                    # 色付ける
                    tgt_ws.cell(row=r, column=c).fill = fill
                    tgt_ws.cell(row=r, column=c + 1).fill = fill
                    r += 1
                c += 2  # 数字抜き出した元の文字列と、抜きだした数字の分
            c += 1


class GetRightNumber3:
    def __init__(self):
        self.name = '右側から数字抽出(適宜挿入)'
        self.explain = '元のデータを書き込みつつ、適宜結果を挿入していく'

    def run(self, workbook, sheet_name, tgt_cols):
        """
        workbook 対象となるワークブック
        sheet name　処理もとのデータがあるシート
        tgt cols　処理を行う列
        """
        src_ws = workbook[sheet_name]  # コピー元のワークシート
        tgt_ws = workbook.create_sheet(title=self.name)  # コピー先のシート
        # tgt_cols を一次元配列に(numpy使いたいけどコード重くなる?)
        tgt_cols_1d = []
        for cols in tgt_cols:
            for col in cols:
                tgt_cols_1d.append(col)
        tgt_cols_1d = set(tgt_cols_1d)
        # 数字抜き出した部分を色付けて付け足す
        matcher = re.compile('\d+$')  # 数字抜き出すための正規表現
        fill = openpyxl.styles.PatternFill(fgColor='FFFF00', bgColor="FFFF00", fill_type="solid")  # 色を塗るためのオブジェクト
        c = 1  # col通し番号
        for rows in src_ws.iter_cols():  # 1列づつ抜き出して処理
            col = openpyxl.utils.get_column_letter(rows[0].column)  # 列のアルファベット
            isTarget = True if col in tgt_cols_1d else False
            r = 1  # 行通し番号
            for row in rows:
                cell = row.value
                if isTarget:  # 処理対象の列か True Falseにしたら毎回判定しても1億回試行して、1.2秒前後の差
                    if cell is not None:
                        result = re.search(matcher, cell.rstrip()) # matchする箇所を探す
                        if result:  # matchした場合
                            tgt_ws.cell(row=r, column=c).value = cell[:result.start()] # match外の部分
                            tgt_ws.cell(row=r, column=c + 1).value = int(result.group()) # matchした部分
                        else:  # matchしない場合そのまま書き込み
                            tgt_ws.cell(row=r, column=c).value = cell
                    # 処理したcellに色をつける
                    tgt_ws.cell(row=r, column=c).fill = fill
                    tgt_ws.cell(row=r, column=c + 1).fill = fill
                else:
                    tgt_ws.cell(row=r, column=c).value = cell  # そのまま書き込み
                r += 1
            c += 1
            if isTarget:  # 処理対象の列だったら実質+2
                c += 1
