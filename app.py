import os
import csv

import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox

import openpyxl
import mojimoji




def create_StringVar(text):
    stringVar = tk.StringVar()
    stringVar.set(text)
    return stringVar


def create_ListBox(frame, functions):
    stringVar = create_StringVar(functions)
    listBox = tk.Listbox(frame, justify='center', selectmode='browse', listvariable=stringVar, height=1)
    scroll_bar = tk.Scrollbar(frame, command=listBox.yview)
    scroll_bar.pack(side='right', fill='y')
    listBox['yscrollcommand'] = scroll_bar.set
    return listBox


def create_ComboBox(frame, variables, state='disabled'):
    comboBox = ttk.Combobox(frame, height=1, width=30, justify='center', state=state, values=variables)
    if len(variables) > 0:
        comboBox.current(0)
    return comboBox


def csv2xlsx(csv_path):
    # csvをエクセルに変換
    excel_path = os.path.join(os.path.dirname(csv_path),
                              '(xlsx)' + os.path.basename(csv_path).replace('.csv', '.xlsx'))  # 変換後のpath
    if not os.path.exists(excel_path):
        # すでにxlsxファイルが存在したらパスをそのまま返し、なければxlsxファイルを作成
        wb = openpyxl.Workbook()  # 新規Excelファイル作成
        ws = wb.active  # 現在のシートをアクティブに
        with open(csv_path) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
        wb.save(excel_path)
    return excel_path



def parse_selected_col(col_string):
    tmp = []
    col_string = mojimoji.zen_to_han(col_string).replace(' ','')
    for char in col_string.split(','):
        if char == '':
            continue
        char_split = char.split(':')
        if len(char_split) == 1:
            tmp.append([char])
        else:
            start = openpyxl.utils.column_index_from_string(char_split[0])
            end = openpyxl.utils.column_index_from_string(char_split[1])
            tmp.append([openpyxl.utils.get_column_letter(i) for i in range(start, end + 1)])
    return tmp


class Application(tk.Tk):
    def __init__(self, funcs):
        super().__init__()
        # 使用する変数
        self.target_file = None
        self.file_type = None  # csv or その他
        self.wb = None  # 読み込んだファイル
        # 使用する関数一覧
        self.funcs = {func.name: func for func in funcs}
        # window の設定
        self.title('数字抜き出し')
        # frameの定義
        self.frame = tk.Frame(self)
        # パーツ作成
        # ファイル選択
        self.label_file = tk.Label(self.frame, text='ファイル選択')
        self.dialog_file = tk.Entry(self.frame, width=40, state='disabled')
        self.btn_select_file = tk.Button(self.frame, text='選択', command=self.select_file)

        # シート選択
        self.label_sheet = tk.Label(self.frame, text='シート選択')
        self.combo_sheet = create_ComboBox(self.frame, [])

        # 列選択
        self.label_col = tk.Label(self.frame, text='列選択')
        self.dialog_col = tk.Entry(self.frame, width=40, state='disabled')
        self.btn_tips_col = tk.Button(self.frame, text='説明', command=self.show_tips)
        # self.btn_select_col = tk.Button(self.frame, text='選択', command=self.select_col, state='disabled')

        # 関数選択
        self.label_func = tk.Label(self.frame, text='操作選択')
        self.combo_func = create_ComboBox(self.frame, list(self.funcs.keys()))

        # 実行
        self.btnRun = tk.Button(self, text='実行', state='disabled', command=self.run)

        # frameに配置
        self.label_file.grid(row=0, column=0, sticky=tk.NSEW, pady=5)
        self.dialog_file.grid(row=0, column=1, sticky=tk.NSEW, pady=5)
        self.btn_select_file.grid(row=0, column=2, sticky=tk.NSEW, pady=5)

        self.label_sheet.grid(row=1, column=0, sticky=tk.NSEW, pady=5)
        self.combo_sheet.grid(row=1, column=1, sticky=tk.NSEW, pady=5)

        self.label_col.grid(row=2, column=0, sticky=tk.NSEW, pady=5)
        self.dialog_col.grid(row=2, column=1, sticky=tk.NSEW, pady=5)
        self.btn_tips_col.grid(row=2, column=2, sticky=tk.NSEW, pady=5)
        # self.btn_select_col.grid(row=2, column=2, sticky=tk.NSEW, pady=5)

        self.label_func.grid(row=3, column=0, sticky=tk.NSEW, pady=5)
        self.combo_func.grid(row=3, column=1, sticky=tk.NSEW, pady=5)

        # windowに配置
        self.frame.pack()
        self.btnRun.pack(pady=5)

    # ファイルの選択と読み込み
    def select_file(self):
        self.dialog_file['state'] = 'normal'  # 変更可能にする
        self.dialog_file.delete(0, tk.END)  # box内を初期化
        self.dialog_file.insert(tk.END,
                                filedialog.askopenfilename(filetypes=[('Excelファイル', '*.csv;*.xlsx;*.xls')]))
        self.dialog_file['state'] = 'readonly'  # 変更不可にする
        # ファイルの種類で処理変える
        self.file_type = self.dialog_file.get().split('.')[-1]
        # 対象ファイルがcsvならxlsxに変換
        if self.file_type == 'csv':
            self.target_file = csv2xlsx(self.dialog_file.get())
        else:
            self.target_file = self.dialog_file.get()
        # ファイル読み込み
        self.wb = openpyxl.load_workbook(self.target_file)
        # シート選択のボタンを有効にする
        sheets = self.wb.sheetnames
        self.combo_sheet['values'] = sheets
        self.combo_sheet.set(sheets[0])
        self.combo_sheet['state'] = 'readonly'
        # 列選択ボタンを有効にする
        self.dialog_col['state'] = 'normal'  # 選択例: A,G:H,AL:AT
        # self.btn_select_col['state'] = 'normal'
        # 操作選択を有効にする
        self.combo_func['state'] = 'normal'
        # 実行ボタンを有効にする
        self.btnRun['state'] = 'normal'

    def select_col(self):
        """
        作成中
        # サブウィンドウ作成
        # 値の入っている列一覧
        cols = [openpyxl.utils.get_column_letter(i + 1) for i in range(self.wb[self.combo_sheet.get()].max_column)]
        # 1行目の内容
        values = [str(self.wb[self.combo_sheet.get()][col][0].value) for col in cols]
        # 列番号+内容
        col_with_value = [f'{cols[i]}({values[i][:12]})' if len(values[i]) > 12 else f'{cols[i]}({values[i]})' for i in
                          range(len(values))]
        """
        print(parse_selected_col(self.dialog_col.get()))

    def run(self):
        tgt_cols = parse_selected_col(self.dialog_col.get())  # 対象となる列一覧
        func_name = self.combo_func.get()  # 適用する関数
        func = self.funcs[func_name]
        tgt_sheet = self.wb.create_sheet(title=f'{func_name}')
        try:
            func.run(self.wb[self.combo_sheet.get()], tgt_sheet, tgt_cols)
            self.wb.save(self.target_file)
            messagebox.showinfo('確認', '変換が完了しました')
        except Exception as e:
            #print('=== エラー内容 ===')
            #print('type:' + str(type(e)))
            #print('args:' + str(e.args))
            #print('message:' + e.message)
            #print('error:' + str(e))
            sentence = '変換中にエラーが発生しました\n' \
                       '=== エラー内容 ===\n' \
                       f'type:{str(type(e))}\n' \
                       f'args:{str(e.args)}\n' \
                       f'message:{e.message}\n' \
                       f'error:{str(e)}'
            messagebox.showerror('警告', sentence)

    def show_tips(self):
        sentence = "列を選択 範囲指定 「:」 複数選択[,]\n" \
                   "例\n" \
                   "A列→A\n" \
                   "C列~H列→C:H\n" \
                   "A列とC列~H列→A,C:H\n"
        messagebox.showinfo('説明', sentence)






